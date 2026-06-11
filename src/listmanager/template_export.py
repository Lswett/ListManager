from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

SUPPORTED_SHEETS = ("FULLNAME", "FIRSTLAST", "COMPANY")

SHEET_HEADERS = {
    "FULLNAME": ["FullName", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
    "FIRSTLAST": ["First Name", "Last Name", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
    "COMPANY": ["ATTN", "Company", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
}


@dataclass(frozen=True)
class TemplateExportResult:
    converted_path: Path
    template_path: Path
    output_path: Path
    source_sheet: str
    target_sheet: str
    rows_exported: int


class TemplateExportError(ValueError):
    pass


def _clean_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def _header_map(values: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(values, 1):
        header = _clean_header(value)
        if header and header not in mapping:
            mapping[header] = idx
    return mapping


def _find_source_sheet(wb) -> str:
    first = wb.sheetnames[0] if wb.sheetnames else ""
    if first in SUPPORTED_SHEETS:
        return first

    matches = [name for name in wb.sheetnames if name in SUPPORTED_SHEETS]
    if not matches:
        raise TemplateExportError(
            "No supported passed-records sheet found. Expected one of: FULLNAME, FIRSTLAST, COMPANY."
        )
    if len(matches) == 1:
        return matches[0]
    raise TemplateExportError(
        "Multiple possible passed-records sheets were found and the first worksheet is not one of them: "
        + ", ".join(matches)
    )


def _missing(required: list[str], mapping: dict[str, int]) -> list[str]:
    return [header for header in required if header not in mapping]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def _zip_text(value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 4:
        return digits.zfill(5)
    if len(digits) == 5:
        return digits
    return text


def _zip4_text(value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if 0 < len(digits) <= 4 and text.replace("-", "").replace(" ", "").isdigit():
        return digits.zfill(4)
    return text


def _row_is_blank(ws, row_idx: int, columns: list[int]) -> bool:
    return all(_cell_text(ws.cell(row_idx, column).value) == "" for column in columns)


def _clear_target_rows(ws, start_row: int = 8) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def export_to_template(converted_path: Path, template_path: Path, output_path: Path) -> TemplateExportResult:
    converted_path = Path(converted_path)
    template_path = Path(template_path)
    output_path = Path(output_path)

    converted_wb = load_workbook(converted_path, data_only=False)
    source_sheet_name = _find_source_sheet(converted_wb)
    source_ws = converted_wb[source_sheet_name]

    if source_sheet_name not in SHEET_HEADERS:
        raise TemplateExportError(f"Unsupported source sheet: {source_sheet_name}")
    required_headers = SHEET_HEADERS[source_sheet_name]

    source_headers = _header_map([source_ws.cell(1, column).value for column in range(1, source_ws.max_column + 1)])
    missing_source = _missing(required_headers, source_headers)
    if missing_source:
        raise TemplateExportError(
            f"Required source headers are missing from source sheet {source_sheet_name} "
            f"for target template tab {source_sheet_name}: {', '.join(missing_source)}"
        )

    template_wb = load_workbook(template_path)
    if source_sheet_name not in template_wb.sheetnames:
        raise TemplateExportError(f"Matching template tab is missing: {source_sheet_name}")
    target_ws = template_wb[source_sheet_name]

    target_headers = _header_map([target_ws.cell(4, column).value for column in range(1, target_ws.max_column + 1)])
    missing_target = _missing(required_headers, target_headers)
    if missing_target:
        raise TemplateExportError(
            f"Required target headers are missing from template tab {source_sheet_name} "
            f"for source sheet {source_sheet_name}: {', '.join(missing_target)}"
        )

    _clear_target_rows(target_ws, start_row=8)

    source_columns = [source_headers[header] for header in required_headers]
    target_row = 8
    rows_exported = 0
    for source_row in range(2, source_ws.max_row + 1):
        if _row_is_blank(source_ws, source_row, source_columns):
            continue
        for header in required_headers:
            value = source_ws.cell(source_row, source_headers[header]).value
            if header == "Zip":
                value = _zip_text(value)
            elif header == "Zip4":
                value = _zip4_text(value)
            else:
                value = _cell_text(value)
            target_cell = target_ws.cell(target_row, target_headers[header])
            target_cell.value = value
            if header in {"Zip", "Zip4"}:
                target_cell.number_format = "@"
        target_row += 1
        rows_exported += 1

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        template_wb.save(output_path)
    except Exception as exc:
        raise TemplateExportError(f"Could not save output workbook {output_path}: {exc}") from exc
    finally:
        converted_wb.close()
        template_wb.close()

    return TemplateExportResult(
        converted_path=converted_path,
        template_path=template_path,
        output_path=output_path,
        source_sheet=source_sheet_name,
        target_sheet=source_sheet_name,
        rows_exported=rows_exported,
    )


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Export passed records from a converted ListManager workbook into the mailing list template."
    )
    parser.add_argument("converted_workbook", help="Converted ListManager workbook")
    parser.add_argument("template_workbook", help="Official mailing list template workbook")
    parser.add_argument("output_workbook", help="Output template-ready workbook")
    args = parser.parse_args(argv)

    result = export_to_template(
        Path(args.converted_workbook),
        Path(args.template_workbook),
        Path(args.output_workbook),
    )
    print(
        f"Exported {result.rows_exported} row(s) from {result.source_sheet} "
        f"to {result.output_path}"
    )


if __name__ == "__main__":
    main()
