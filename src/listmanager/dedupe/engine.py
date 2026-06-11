from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

SUPPORTED_SHEETS = ("FULLNAME", "FIRSTLAST", "COMPANY")

COLUMN_ALIASES = {
    "unique_id": ("UniqueFileID", "UniqueFile", "ID", "Record ID"),
    "full_name": ("FullName", "Full Name", "Name"),
    "first_name": ("First Name", "FirstName", "First"),
    "last_name": ("Last Name", "LastName", "Last"),
    "attn": ("ATTN", "Attention"),
    "company": ("Company", "Organization", "Org"),
    "primary_address": ("PrimaryAddress", "Primary Address", "PrimaryAdd", "Address", "Address1"),
    "city": ("City",),
    "state": ("State",),
    "zip": ("Zip", "ZIP", "Zip5", "ZIP5", "PostalCode", "Postal Code"),
}

REQUIRED_FIELDS = ("primary_address", "city", "state", "zip")
MATCHING_RULES = (
    "Exact Individual Duplicate: normalized full name or first+last, primary address, city, state, ZIP5.",
    "Exact Household Duplicate / Possible Duplicate: normalized last name, primary address, city, state, ZIP5.",
    "Address-only Possible Duplicate: normalized primary address, city, state, ZIP5.",
)

HELPER_REMOVED_HEADERS = [
    "Duplicate Group ID",
    "Duplicate Type",
    "Duplicate Reason",
    "Kept Record Row/ID",
]
HELPER_REVIEW_HEADERS = [
    "Duplicate Group ID",
    "Duplicate Type",
    "Duplicate Reason",
    "Suggested Action",
]


class DedupeError(ValueError):
    pass


@dataclass(frozen=True)
class DedupeOptions:
    auto_remove_exact_individuals: bool = True
    create_possible_review: bool = True


@dataclass(frozen=True)
class DedupeResult:
    input_path: Path
    output_dir: Path
    data_sheet: str
    input_records: int
    output_records: int
    removed_duplicates: int
    possible_duplicate_groups: int
    possible_duplicate_records: int
    deduped_output_path: Path
    removed_duplicates_path: Path
    possible_duplicates_path: Path
    report_path: Path


@dataclass
class _Record:
    values: dict[str, object]
    excel_row: int
    index: int
    non_empty_count: int


@dataclass(frozen=True)
class _SheetLayout:
    sheet_name: str
    header_row: int
    data_start_row: int
    headers: list[str]
    columns: dict[str, int]
    fields: dict[str, str]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def _header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", _cell_text(value).lower())


def _normalize_text(value: object) -> str:
    text = _cell_text(value).upper()
    text = re.sub(r"[^\w\s#]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_zip(value: object) -> str:
    digits = re.sub(r"\D", "", _cell_text(value))
    if len(digits) >= 5:
        return digits[:5]
    return digits


def _normalize_address(value: object) -> str:
    text = _normalize_text(value)
    replacements = {
        "STREET": "ST",
        "ST": "ST",
        "ROAD": "RD",
        "RD": "RD",
        "AVENUE": "AVE",
        "AVE": "AVE",
        "DRIVE": "DR",
        "DR": "DR",
        "LANE": "LN",
        "LN": "LN",
        "BOULEVARD": "BLVD",
        "BLVD": "BLVD",
        "COURT": "CT",
        "CT": "CT",
        "APARTMENT": "APT",
        "APT": "APT",
        "#": "APT",
    }
    parts = [replacements.get(part, part) for part in text.split()]
    return " ".join(parts)


def _full_name(record: _Record, fields: dict[str, str]) -> str:
    if fields.get("full_name"):
        value = _normalize_text(record.values.get(fields["full_name"], ""))
        if value:
            return value
    first = _normalize_text(record.values.get(fields.get("first_name", ""), ""))
    last = _normalize_text(record.values.get(fields.get("last_name", ""), ""))
    first_last = " ".join(part for part in (first, last) if part)
    if first_last:
        return first_last
    for field in ("attn", "company"):
        if fields.get(field):
            value = _normalize_text(record.values.get(fields[field], ""))
            if value:
                return value
    return ""


def _last_name(record: _Record, fields: dict[str, str]) -> str:
    if fields.get("last_name"):
        return _normalize_text(record.values.get(fields["last_name"], ""))
    name = _full_name(record, fields)
    return name.split()[-1] if name else ""


def _address_key(record: _Record, fields: dict[str, str]) -> tuple[str, str, str, str]:
    return (
        _normalize_address(record.values.get(fields["primary_address"], "")),
        _normalize_text(record.values.get(fields["city"], "")),
        _normalize_text(record.values.get(fields["state"], "")),
        _normalize_zip(record.values.get(fields["zip"], "")),
    )


def _exact_key(record: _Record, fields: dict[str, str]) -> tuple[str, str, str, str, str] | None:
    name = _full_name(record, fields)
    address = _address_key(record, fields)
    if not name or not all(address):
        return None
    return (name, *address)


def _household_key(record: _Record, fields: dict[str, str]) -> tuple[str, str, str, str, str] | None:
    last = _last_name(record, fields)
    address = _address_key(record, fields)
    if not last or not all(address):
        return None
    return (last, *address)


def _possible_address_key(record: _Record, fields: dict[str, str]) -> tuple[str, str, str, str] | None:
    address = _address_key(record, fields)
    if not all(address):
        return None
    return address


def _match_alias(headers: Iterable[str], aliases: Iterable[str]) -> str:
    by_key = {_header_key(header): header for header in headers if _cell_text(header)}
    for alias in aliases:
        found = by_key.get(_header_key(alias))
        if found:
            return found
    return ""


def _field_map(headers: list[str]) -> dict[str, str]:
    return {field: _match_alias(headers, aliases) for field, aliases in COLUMN_ALIASES.items()}


def _missing_required(fields: dict[str, str]) -> list[str]:
    missing = [field for field in REQUIRED_FIELDS if not fields.get(field)]
    if not any(
        (
            fields.get("full_name"),
            fields.get("first_name") and fields.get("last_name"),
            fields.get("attn"),
            fields.get("company"),
        )
    ):
        missing.append("full_name, first_name+last_name, attn, or company")
    return missing


def _header_values(ws, row: int) -> list[str]:
    return [_cell_text(ws.cell(row, column).value) for column in range(1, ws.max_column + 1)]


def _find_sheet_layout(wb) -> _SheetLayout:
    candidates = [name for name in wb.sheetnames if name in SUPPORTED_SHEETS] or list(wb.sheetnames[:1])
    errors: list[str] = []
    for sheet_name in candidates:
        ws = wb[sheet_name]
        for header_row, data_start_row in ((4, 8), (1, 2)):
            headers = _header_values(ws, header_row)
            fields = _field_map(headers)
            missing = _missing_required(fields)
            if not missing:
                columns = {header: idx for idx, header in enumerate(headers, 1) if header}
                return _SheetLayout(sheet_name, header_row, data_start_row, headers, columns, fields)
            errors.append(f"{sheet_name} row {header_row}: missing {', '.join(missing)}")
    raise DedupeError(
        "Could not find a supported List Manager data sheet with required duplicate matching columns. "
        + "; ".join(errors)
    )


def _read_records(ws, layout: _SheetLayout) -> list[_Record]:
    source_columns = [layout.columns[header] for header in layout.headers if header]
    records: list[_Record] = []
    for row_idx in range(layout.data_start_row, ws.max_row + 1):
        values = {header: ws.cell(row_idx, layout.columns[header]).value for header in layout.headers if header}
        if all(_cell_text(ws.cell(row_idx, column).value) == "" for column in source_columns):
            continue
        records.append(
            _Record(
                values=values,
                excel_row=row_idx,
                index=len(records),
                non_empty_count=sum(1 for value in values.values() if _cell_text(value)),
            )
        )
    return records


def _group(records: Iterable[_Record], key_func) -> dict[tuple[object, ...], list[_Record]]:
    groups: dict[tuple[object, ...], list[_Record]] = {}
    for record in records:
        key = key_func(record)
        if key is None:
            continue
        groups.setdefault(key, []).append(record)
    return {key: group for key, group in groups.items() if len(group) > 1}


def _choose_best(records: list[_Record]) -> _Record:
    return max(records, key=lambda record: (record.non_empty_count, -record.index))


def _copy_cell(source, target) -> None:
    target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def _write_records_workbook(path: Path, headers: list[str], rows: list[dict[str, object]], helper_headers: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"
    all_headers = headers + helper_headers
    ws.append(all_headers)
    for row in rows:
        ws.append([row.get(header, "") for header in all_headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _write_report(path: Path, result_values: dict[str, object]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate Report"
    ws.append(["Metric", "Value"])
    for key, value in result_values.items():
        ws.append([key, value])
    ws.append([])
    ws.append(["Matching Rules Used", ""])
    for rule in MATCHING_RULES:
        ws.append(["", rule])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _unique_records(records: Iterable[_Record]) -> list[_Record]:
    seen: set[int] = set()
    unique: list[_Record] = []
    for record in records:
        if record.index in seen:
            continue
        seen.add(record.index)
        unique.append(record)
    return unique


def remove_duplicates(
    input_path: Path,
    output_dir: Path | None = None,
    options: DedupeOptions | None = None,
) -> DedupeResult:
    input_path = Path(input_path)
    if input_path.suffix.lower() != ".xlsx":
        raise DedupeError("Remove Duplicates expects an .xlsx List Manager template-format workbook.")
    if not input_path.is_file():
        raise DedupeError(f"Input workbook not found: {input_path}")

    options = options or DedupeOptions()
    output_dir = Path(output_dir) if output_dir else input_path.with_name(f"{input_path.stem}_dedupe")
    deduped_output_path = output_dir / "deduped_output.xlsx"
    removed_duplicates_path = output_dir / "removed_duplicates.xlsx"
    possible_duplicates_path = output_dir / "possible_duplicates_review.xlsx"
    report_path = output_dir / "duplicate_report.xlsx"

    wb = load_workbook(input_path)
    layout = _find_sheet_layout(wb)
    ws = wb[layout.sheet_name]
    records = _read_records(ws, layout)

    exact_groups = _group(records, lambda record: _exact_key(record, layout.fields))
    removed: list[dict[str, object]] = []
    removed_indexes: set[int] = set()
    exact_group_number = 1
    if options.auto_remove_exact_individuals:
        for group_records in exact_groups.values():
            kept = _choose_best(group_records)
            group_id = f"EXACT-{exact_group_number:04d}"
            exact_group_number += 1
            kept_id = kept.values.get(layout.fields.get("unique_id", ""), "") or kept.excel_row
            for record in group_records:
                if record.index == kept.index:
                    continue
                removed_indexes.add(record.index)
                row = dict(record.values)
                row.update(
                    {
                        "Duplicate Group ID": group_id,
                        "Duplicate Type": "Exact Individual Duplicate",
                        "Duplicate Reason": "Same normalized name, primary address, city, state, and ZIP5.",
                        "Kept Record Row/ID": kept_id,
                    }
                )
                removed.append(row)

    survivors = [record for record in records if record.index not in removed_indexes]
    review_rows: list[dict[str, object]] = []
    review_record_indexes: set[int] = set()
    review_group_count = 0
    if options.create_possible_review:
        household_groups = _group(survivors, lambda record: _household_key(record, layout.fields))
        for group_records in household_groups.values():
            group_id = f"HOUSEHOLD-{review_group_count + 1:04d}"
            review_group_count += 1
            for record in group_records:
                review_record_indexes.add(record.index)
                row = dict(record.values)
                row.update(
                    {
                        "Duplicate Group ID": group_id,
                        "Duplicate Type": "Exact Household Duplicate / Possible Duplicate",
                        "Duplicate Reason": "Same normalized last name and mailing address.",
                        "Suggested Action": "Review manually; not auto-removed.",
                    }
                )
                review_rows.append(row)

        address_groups = _group(survivors, lambda record: _possible_address_key(record, layout.fields))
        for group_records in address_groups.values():
            unique_last_names = {_last_name(record, layout.fields) for record in group_records}
            if len(unique_last_names) <= 1:
                continue
            group_id = f"ADDRESS-{review_group_count + 1:04d}"
            review_group_count += 1
            for record in group_records:
                review_record_indexes.add(record.index)
                row = dict(record.values)
                row.update(
                    {
                        "Duplicate Group ID": group_id,
                        "Duplicate Type": "Address-only Possible Duplicate",
                        "Duplicate Reason": "Same normalized mailing address with different names.",
                        "Suggested Action": "Review manually; not auto-removed.",
                    }
                )
                review_rows.append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    for record in sorted((r for r in records if r.index in removed_indexes), key=lambda r: r.excel_row, reverse=True):
        ws.delete_rows(record.excel_row, 1)
    wb.save(deduped_output_path)
    wb.close()

    base_headers = [header for header in layout.headers if header]
    _write_records_workbook(removed_duplicates_path, base_headers, removed, HELPER_REMOVED_HEADERS)
    _write_records_workbook(possible_duplicates_path, base_headers, review_rows, HELPER_REVIEW_HEADERS)

    result = DedupeResult(
        input_path=input_path,
        output_dir=output_dir,
        data_sheet=layout.sheet_name,
        input_records=len(records),
        output_records=len(records) - len(removed_indexes),
        removed_duplicates=len(removed_indexes),
        possible_duplicate_groups=review_group_count,
        possible_duplicate_records=len(_unique_records(record for record in survivors if record.index in review_record_indexes)),
        deduped_output_path=deduped_output_path,
        removed_duplicates_path=removed_duplicates_path,
        possible_duplicates_path=possible_duplicates_path,
        report_path=report_path,
    )
    _write_report(
        report_path,
        {
            "Input filename": input_path.name,
            "Data sheet": result.data_sheet,
            "Timestamp": datetime.now().isoformat(timespec="seconds"),
            "Input record count": result.input_records,
            "Output record count": result.output_records,
            "Auto-removed duplicate count": result.removed_duplicates,
            "Possible duplicate group count": result.possible_duplicate_groups,
            "Possible duplicate record count": result.possible_duplicate_records,
        },
    )
    return result
