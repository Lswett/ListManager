from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .mapper import detect_target_format, header_score, map_headers

_ZIPISH_RE = re.compile(r"^\s*\d{4,5}(?:\.0)?(?:-\d{4})?\s*$")
_STREET_RE = re.compile(
    r"\b(\d+|p\.?\s*o\.?\s*box|po box|street|st|road|rd|avenue|ave|drive|dr|lane|ln|blvd|way|ct|court)\b",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class SourceRow:
    source_row: int
    values: list[object]
    original: dict[str, object]


@dataclass(frozen=True)
class WorkbookScan:
    source_path: Path
    sheet_name: str
    target_format: str
    header_row: int | None
    headers: list[str]
    field_map: dict[str, int]
    rows: list[SourceRow]
    headerless: bool = False
    assumptions: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()


def _row_values(ws: Worksheet, row_number: int) -> list[object]:
    return [cell.value for cell in ws[row_number]]


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _nonempty_count(values: list[object]) -> int:
    return sum(1 for value in values if _text(value))


def _looks_headerless(values: list[object]) -> bool:
    if len(values) < 5:
        return False
    first = _text(values[0])
    address = _text(values[2])
    city = _text(values[3])
    zip_value = _text(values[4])
    return bool("," in first and _STREET_RE.search(address) and city and _ZIPISH_RE.match(zip_value))


def _best_header_row(ws: Worksheet) -> tuple[int | None, int]:
    best_row: int | None = None
    best_score = 0
    for row_number in range(1, min(ws.max_row, 25) + 1):
        values = _row_values(ws, row_number)
        score = header_score(values)
        if score > best_score:
            best_row = row_number
            best_score = score
    if best_score >= 2:
        return best_row, best_score
    return None, 0


def _sheet_score(ws: Worksheet) -> tuple[int, bool, int | None]:
    header_row, score = _best_header_row(ws)
    if header_row:
        data_rows = sum(
            1 for row_number in range(header_row + 1, ws.max_row + 1)
            if _nonempty_count(_row_values(ws, row_number)) >= 2
        )
        return score * 10 + data_rows, False, header_row

    headerless_rows = sum(
        1 for row_number in range(1, min(ws.max_row, 20) + 1)
        if _looks_headerless(_row_values(ws, row_number))
    )
    return headerless_rows * 20, bool(headerless_rows), None


def _original(headers: list[str], values: list[object]) -> dict[str, object]:
    original: dict[str, object] = {}
    for idx, header in enumerate(headers):
        value = values[idx] if idx < len(values) else ""
        original[header] = value
    return original


def scan_workbook(path: Path) -> WorkbookScan:
    wb = load_workbook(path, data_only=True)
    scored = [(_sheet_score(ws), ws) for ws in wb.worksheets]
    scored.sort(key=lambda item: item[0][0], reverse=True)
    (_, headerless, header_row), ws = scored[0]

    if headerless:
        headers = ["LastFirst", "Grade", "PrimaryAddress", "City", "Zip"]
        field_map = {"First Name": 0, "Last Name": 0, "PrimaryAddress": 2, "City": 3, "Zip": 4}
        target_format, warnings = detect_target_format(field_map, headerless=True)
        rows: list[SourceRow] = []
        for row_number in range(1, ws.max_row + 1):
            values = _row_values(ws, row_number)
            if _nonempty_count(values) == 0:
                continue
            rows.append(SourceRow(row_number, values, _original(headers, values)))
        return WorkbookScan(
            source_path=path,
            sheet_name=ws.title,
            target_format=target_format,
            header_row=None,
            headers=headers,
            field_map=field_map,
            rows=rows,
            headerless=True,
            assumptions=("Headerless Last, First | Grade | Address | City | Zip pattern detected.",),
            warnings=tuple(warnings),
        )

    if header_row is None:
        ws = wb.worksheets[0]
        headers = [f"Column {idx}" for idx in range(1, ws.max_column + 1)]
        rows = [
            SourceRow(row_number, _row_values(ws, row_number), _original(headers, _row_values(ws, row_number)))
            for row_number in range(1, ws.max_row + 1)
            if _nonempty_count(_row_values(ws, row_number)) > 0
        ]
        return WorkbookScan(path, ws.title, "UNKNOWN", None, headers, {}, rows)

    headers = [_text(value) or f"Column {idx + 1}" for idx, value in enumerate(_row_values(ws, header_row))]
    field_map = map_headers(headers)
    target_format, warnings = detect_target_format(field_map)
    rows = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        values = _row_values(ws, row_number)
        if _nonempty_count(values) == 0:
            continue
        rows.append(SourceRow(row_number, values, _original(headers, values)))
    return WorkbookScan(path, ws.title, target_format, header_row, headers, field_map, rows, warnings=tuple(warnings))
