from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .mapper import REVIEW_COLUMNS, TARGET_COLUMNS


def _write_sheet(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        values = [row.get(header, "") for header in headers]
        ws.append(values)
    for column_name in ("Zip", "Zip4"):
        if column_name in headers:
            idx = headers.index(column_name) + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = "@"


def write_converted_workbook(
    output_path: Path,
    target_format: str,
    main_rows: list[dict[str, object]],
    review_rows: list[dict[str, object]],
    report_rows: list[tuple[str, object]],
    original_headers: list[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = target_format if target_format in TARGET_COLUMNS else "NEEDS_REVIEW"

    if target_format in TARGET_COLUMNS:
        _write_sheet(ws, TARGET_COLUMNS[target_format], main_rows)
        review_ws = wb.create_sheet("NEEDS_REVIEW")
    else:
        review_ws = ws
        review_ws.title = "NEEDS_REVIEW"

    review_headers = REVIEW_COLUMNS + [f"ORIGINAL: {header}" for header in original_headers]
    _write_sheet(review_ws, review_headers, review_rows)

    report_ws = wb.create_sheet("CONVERSION_REPORT")
    report_ws.append(["Metric", "Value"])
    for cell in report_ws[1]:
        cell.font = Font(bold=True)
    for key, value in report_rows:
        report_ws.append([key, value])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
