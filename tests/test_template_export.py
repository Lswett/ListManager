from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from listmanager.template_export import TemplateExportError, export_to_template


SHEET_HEADERS = {
    "COMPANY": ["ATTN", "Company", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
    "FULLNAME": ["FullName", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
    "FIRSTLAST": ["First Name", "Last Name", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"],
}


def _create_template(path: Path, omit_sheet: str | None = None, omit_header: str | None = None) -> None:
    wb = Workbook()
    start = wb.active
    start.title = "START HERE"
    start["A1"] = "Mailing List Template v4"
    start["A7"] = "Instructions stay here"

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name == omit_sheet:
            continue
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{sheet_name} Mailing List"
        ws["A2"] = "Paste your data starting on ROW 8"
        ws["A3"] = "Required columns are shaded gold."
        for col_idx, header in enumerate(headers, 1):
            ws.cell(4, col_idx).value = "" if header == omit_header else header
        ws["A5"] = "EXAMPLES"
        ws["A6"] = "Example value"
        ws["A7"] = "Do not overwrite"
        ws["A8"] = "old pasted data"
    wb.save(path)


def _create_converted(path: Path, sheet_name: str, rows: list[list[object]], headers: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers or SHEET_HEADERS[sheet_name])
    for row in rows:
        ws.append(row)
    review = wb.create_sheet("NEEDS_REVIEW")
    review.append(["Error Codes", "FullName"])
    review.append(["ZIP_INVALID", "Should Not Copy"])
    report = wb.create_sheet("CONVERSION_REPORT")
    report.append(["Metric", "Value"])
    report.append(["Rows moved to NEEDS_REVIEW", 1])
    wb.save(path)


def _load(path: Path):
    return load_workbook(path, data_only=False)


class TemplateExportTests(unittest.TestCase):
    def test_fullname_exports_to_row_8(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "0123", "US"]],
            )

            result = export_to_template(converted, template, output)
            wb = _load(output)

            self.assertEqual(result.rows_exported, 1)
            self.assertEqual(wb["FULLNAME"]["A8"].value, "Jane Smith")
            self.assertEqual(wb["FULLNAME"]["F8"].value, "06110")
            self.assertEqual(wb["FULLNAME"]["G8"].value, "0123")
            wb.close()

    def test_firstlast_exports_to_row_8(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FIRSTLAST",
                [["Jane", "Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)

            self.assertEqual(wb["FIRSTLAST"]["A8"].value, "Jane")
            self.assertEqual(wb["FIRSTLAST"]["B8"].value, "Smith")
            wb.close()

    def test_company_exports_to_row_8(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "COMPANY",
                [["Director", "Acme School", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)

            self.assertEqual(wb["COMPANY"]["A8"].value, "Director")
            self.assertEqual(wb["COMPANY"]["B8"].value, "Acme School")
            wb.close()

    def test_template_instructions_and_headers_are_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)

            self.assertEqual(wb["FULLNAME"]["A1"].value, "FULLNAME Mailing List")
            self.assertEqual(wb["FULLNAME"]["A7"].value, "Do not overwrite")
            self.assertEqual([wb["FULLNAME"].cell(4, c).value for c in range(1, 9)], SHEET_HEADERS["FULLNAME"])
            wb.close()

    def test_review_and_report_are_not_copied(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)
            values = [cell.value for row in wb["FULLNAME"].iter_rows() for cell in row]

            self.assertNotIn("Should Not Copy", values)
            self.assertNotIn("Rows moved to NEEDS_REVIEW", values)
            wb.close()

    def test_zip_and_zip4_remain_text_and_preserve_leading_zeroes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", 6110, 123, "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)
            zip_cell = wb["FULLNAME"]["F8"]
            zip4_cell = wb["FULLNAME"]["G8"]

            self.assertEqual(zip_cell.value, "06110")
            self.assertEqual(zip4_cell.value, "0123")
            self.assertEqual(zip_cell.number_format, "@")
            self.assertEqual(zip4_cell.number_format, "@")
            wb.close()

    def test_other_template_tabs_still_exist_and_are_not_damaged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)

            self.assertEqual(wb.sheetnames, ["START HERE", "COMPANY", "FULLNAME", "FIRSTLAST"])
            self.assertEqual(wb["COMPANY"]["A8"].value, "old pasted data")
            self.assertEqual(wb["START HERE"]["A1"].value, "Mailing List Template v4")
            wb.close()

    def test_output_workbook_opens_successfully(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            converted = root / "converted.xlsx"
            output = root / "output.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            export_to_template(converted, template, output)
            wb = _load(output)
            self.assertIn("FULLNAME", wb.sheetnames)
            wb.close()

    def test_fails_if_no_supported_source_sheet_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            converted = root / "converted.xlsx"
            template = root / "template.xlsx"
            _create_template(template)
            wb = Workbook()
            wb.active.title = "NEEDS_REVIEW"
            wb.save(converted)

            with self.assertRaisesRegex(TemplateExportError, "No supported passed-records sheet"):
                export_to_template(converted, template, root / "out.xlsx")

    def test_fails_if_matching_template_tab_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            converted = root / "converted.xlsx"
            template = root / "template.xlsx"
            _create_template(template, omit_sheet="FULLNAME")
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            with self.assertRaisesRegex(TemplateExportError, "Matching template tab is missing"):
                export_to_template(converted, template, root / "out.xlsx")

    def test_fails_if_source_headers_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            converted = root / "converted.xlsx"
            template = root / "template.xlsx"
            _create_template(template)
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St"]],
                headers=["FullName", "PrimaryAddress"],
            )

            with self.assertRaisesRegex(TemplateExportError, "Required source headers are missing"):
                export_to_template(converted, template, root / "out.xlsx")

    def test_fails_if_target_headers_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            converted = root / "converted.xlsx"
            template = root / "template.xlsx"
            _create_template(template, omit_header="Zip")
            _create_converted(
                converted,
                "FULLNAME",
                [["Jane Smith", "1 Main St", "", "West Hartford", "CT", "06110", "", "US"]],
            )

            with self.assertRaisesRegex(TemplateExportError, "Required target headers are missing"):
                export_to_template(converted, template, root / "out.xlsx")


if __name__ == "__main__":
    unittest.main()
