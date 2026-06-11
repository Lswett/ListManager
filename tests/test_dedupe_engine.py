from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from listmanager.dedupe import DedupeError, remove_duplicates


HEADERS = ["FullName", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country", "Email"]


def _create_template_workbook(path: Path, rows: list[list[object]], headers: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FULLNAME"
    for col_idx, header in enumerate(headers or HEADERS, 1):
        ws.cell(4, col_idx).value = header
    for row_idx, row in enumerate(rows, 8):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx).value = value
    wb.save(path)
    wb.close()


def _create_company_workbook(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "COMPANY"
    headers = ["ATTN", "Company", "PrimaryAddress", "Address2", "City", "State", "Zip", "Zip4", "Country"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(4, col_idx).value = header
    for row_idx, row in enumerate(rows, 8):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx).value = value
    wb.save(path)
    wb.close()


def _sheet_rows(path: Path, sheet_name: str = "FULLNAME") -> list[list[object]]:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    rows = [
        [ws.cell(row_idx, col_idx).value for col_idx in range(1, len(HEADERS) + 1)]
        for row_idx in range(8, ws.max_row + 1)
        if any(ws.cell(row_idx, col_idx).value for col_idx in range(1, len(HEADERS) + 1))
    ]
    wb.close()
    return rows


class DedupeEngineTests(unittest.TestCase):
    def test_exact_duplicate_same_name_and_address_is_removed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.input_records, 2)
            self.assertEqual(result.output_records, 1)
            self.assertEqual(result.removed_duplicates, 1)
            self.assertEqual(len(_sheet_rows(result.deduped_output_path)), 1)

    def test_case_spacing_punctuation_and_zip4_are_normalized(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane   Smith", "1 Main Street Apt. 2", "", "Lansing", "MI", "48910-1234", "", "US", ""],
                    [" jane smith ", "1 MAIN ST # 2", "", " lansing ", "mi", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.removed_duplicates, 1)
            self.assertEqual(result.output_records, 1)

    def test_same_household_goes_to_review_not_removed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["John Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.removed_duplicates, 0)
            self.assertEqual(result.output_records, 2)
            self.assertEqual(result.possible_duplicate_groups, 1)
            self.assertEqual(result.possible_duplicate_records, 2)

    def test_same_address_different_last_names_goes_to_review_not_removed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["John Jones", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.removed_duplicates, 0)
            self.assertEqual(result.possible_duplicate_groups, 1)
            self.assertEqual(result.possible_duplicate_records, 2)

    def test_clearly_different_addresses_are_not_flagged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["Jane Smith", "10 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.removed_duplicates, 0)
            self.assertEqual(result.possible_duplicate_groups, 0)

    def test_missing_optional_fields_do_not_crash(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            headers = ["FullName", "PrimaryAddress", "City", "State", "Zip"]
            _create_template_workbook(
                source,
                [["Jane Smith", "1 Main St", "Lansing", "MI", "48910"]],
                headers=headers,
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.input_records, 1)
            self.assertEqual(result.output_records, 1)

    def test_missing_required_fields_raise_clear_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [["Jane Smith", "Lansing", "MI", "48910"]],
                headers=["FullName", "City", "State", "Zip"],
            )

            with self.assertRaisesRegex(DedupeError, "required duplicate matching columns"):
                remove_duplicates(source, root / "out")

    def test_best_record_selection_keeps_more_populated_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["Jane Smith", "1 Main St", "Apt 2", "Lansing", "MI", "48910", "", "US", "jane@example.com"],
                ],
            )

            result = remove_duplicates(source, root / "out")
            rows = _sheet_rows(result.deduped_output_path)

            self.assertEqual(result.removed_duplicates, 1)
            self.assertEqual(rows[0][2], "Apt 2")
            self.assertEqual(rows[0][8], "jane@example.com")

    def test_output_artifacts_and_counts_are_written(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "input.xlsx"
            _create_template_workbook(
                source,
                [
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["Jane Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                    ["John Smith", "1 Main St", "", "Lansing", "MI", "48910", "", "US", ""],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertTrue(result.deduped_output_path.exists())
            self.assertTrue(result.removed_duplicates_path.exists())
            self.assertTrue(result.possible_duplicates_path.exists())
            self.assertTrue(result.report_path.exists())
            self.assertEqual(result.input_records, 3)
            self.assertEqual(result.output_records, 2)
            self.assertEqual(result.removed_duplicates, 1)
            self.assertEqual(result.possible_duplicate_groups, 1)
            self.assertEqual(result.possible_duplicate_records, 2)

    def test_company_template_sheet_uses_company_identity_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "company.xlsx"
            _create_company_workbook(
                source,
                [
                    ["Director", "Acme School", "1 Main St", "", "Lansing", "MI", "48910", "", "US"],
                    ["Director", "Acme School", "1 Main St", "", "Lansing", "MI", "48910", "", "US"],
                ],
            )

            result = remove_duplicates(source, root / "out")

            self.assertEqual(result.data_sheet, "COMPANY")
            self.assertEqual(result.removed_duplicates, 1)
            self.assertEqual(result.output_records, 1)


if __name__ == "__main__":
    unittest.main()
