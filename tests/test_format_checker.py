from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from listmanager.format_checker import convert_workbook


def _make_workbook(path: Path, headers: list[str] | None, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    if headers is not None:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _convert(path: Path, output_dir: Path):
    result = convert_workbook(path, output_dir)
    assert result.output_path is not None
    return result, load_workbook(result.output_path)


class FormatCheckerTests(unittest.TestCase):
    def test_company_detection_wins_over_firstlast(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "company.xlsx"
            _make_workbook(
                source,
                ["Organization", "First Name", "Last Name", "Street Address", "City", "State", "Zip"],
                [["Acme School", "Jane", "Smith", "1 Main St", "West Hartford", "CT", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.detected_format, "COMPANY")
            self.assertIn("COMPANY", wb.sheetnames)
            self.assertEqual(wb["COMPANY"]["A2"].value, "Acme School")
            self.assertEqual(wb["COMPANY"]["B2"].value, "Jane Smith")

    def test_firstlast_detection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "firstlast.xlsx"
            _make_workbook(
                source,
                ["First", "Last", "Address", "City", "State", "Zip"],
                [["Jane", "Smith", "1 Main St", "West Hartford", "CT", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.detected_format, "FIRSTLAST")
            self.assertIn("FIRSTLAST", wb.sheetnames)

    def test_fullname_detection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "fullname.xlsx"
            _make_workbook(
                source,
                ["Student Name", "Address", "City", "State", "Zip"],
                [["Jane Smith", "1 Main St", "West Hartford", "CT", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.detected_format, "FULLNAME")
            self.assertIn("FULLNAME", wb.sheetnames)

    def test_headerless_last_first_converts_to_firstlast(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "headerless.xlsx"
            _make_workbook(
                source,
                None,
                [["Smith, Jane M", "12", "1 Main St", "West Hartford", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.detected_format, "FIRSTLAST")
            self.assertEqual(wb["FIRSTLAST"]["A2"].value, "Jane")
            self.assertEqual(wb["FIRSTLAST"]["B2"].value, "Smith")
            self.assertEqual(wb["FIRSTLAST"]["F2"].value, "CT")
            report_values = [cell.value for row in wb["CONVERSION_REPORT"].iter_rows(values_only=False) for cell in row]
            self.assertIn("NAME_SPLIT_LAST_FIRST: 1; STATE_FILLED_FROM_ZIP: 1", report_values)

    def test_zip_padding_preserves_leading_zero(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "zip_pad.xlsx"
            _make_workbook(
                source,
                ["First", "Last", "Address", "City", "Zip"],
                [["Jane", "Smith", "1 Main St", "West Hartford", "6110"]],
            )
            _, wb = _convert(source, root / "out")

            self.assertEqual(wb["FIRSTLAST"]["G2"].value, "06110")

    def test_zip4_splits(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "zip4.xlsx"
            _make_workbook(
                source,
                ["First", "Last", "Address", "City", "State", "Zip"],
                [["Jane", "Smith", "1 Main St", "Flint", "MI", "48504-8400"]],
            )
            _, wb = _convert(source, root / "out")

            self.assertEqual(wb["FIRSTLAST"]["G2"].value, "48504")
            self.assertEqual(wb["FIRSTLAST"]["H2"].value, "8400")

    def test_missing_state_is_filled_from_zip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "state_fill.xlsx"
            _make_workbook(
                source,
                ["First", "Last", "Address", "City", "Zip"],
                [["Jane", "Smith", "1 Main St", "West Hartford", "06110"]],
            )
            _, wb = _convert(source, root / "out")

            self.assertEqual(wb["FIRSTLAST"]["F2"].value, "CT")

    def test_state_zip_mismatch_goes_to_needs_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "mismatch.xlsx"
            _make_workbook(
                source,
                ["First", "Last", "Address", "City", "State", "Zip"],
                [["Jane", "Smith", "1 Main St", "West Hartford", "MI", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.rows_needing_review, 1)
            self.assertIn("STATE_ZIP_MISMATCH", wb["NEEDS_REVIEW"]["E2"].value)

    def test_international_country_goes_to_needs_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "intl.xlsx"
            _make_workbook(
                source,
                ["Full Name", "Address", "City", "Province", "Postal Code", "Country"],
                [["Jane Smith", "10 King St", "Ottawa", "ON", "K1A 0B1", "Canada"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.rows_needing_review, 1)
            self.assertIn("INTERNATIONAL_MAIL_REVIEW_REQUIRED", wb["NEEDS_REVIEW"]["E2"].value)

    def test_missing_address_goes_to_needs_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "missing_address.xlsx"
            _make_workbook(
                source,
                ["Full Name", "Address", "City", "State", "Zip"],
                [["Jane Smith", "", "West Hartford", "CT", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.rows_needing_review, 1)
            self.assertIn("PRIMARY_ADDRESS_MISSING", wb["NEEDS_REVIEW"]["E2"].value)

    def test_converted_workbook_contains_required_sheets_and_review_same_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "sheets.xlsx"
            _make_workbook(
                source,
                ["Full Name", "Address", "City", "State", "Zip"],
                [["Jane Smith", "1 Main St", "West Hartford", "CT", "06110"]],
            )
            result, wb = _convert(source, root / "out")

            self.assertEqual(set(wb.sheetnames), {"FULLNAME", "NEEDS_REVIEW", "CONVERSION_REPORT"})
            self.assertTrue(result.output_path and result.output_path.exists())
            self.assertFalse((root / "out" / "sheets_needs_review.xlsx").exists())

    def test_duplicates_are_not_removed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "dupes.xlsx"
            row = ["Jane Smith", "1 Main St", "West Hartford", "CT", "06110"]
            _make_workbook(source, ["Full Name", "Address", "City", "State", "Zip"], [row, row])
            result, wb = _convert(source, root / "out")

            self.assertEqual(result.rows_converted, 2)
            self.assertEqual(wb["FULLNAME"].max_row, 3)


if __name__ == "__main__":
    unittest.main()
